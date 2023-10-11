# -*- coding:utf-8 -*-
"""
Date: 2023-09-18 22:00:56
Desc: 均线-布林带买卖回测
"""

import os
import traceback
import pandas as pd
from datetime import datetime
from openpyxl import Workbook,load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import Reference,ScatterChart  # Reference：图标所用信息
from openpyxl.chart import Series
import trading.collector.constant as ccons
import trading.strategy.calc as calc
import trading.strategy.constant as scons
from trading.config.logger import logger

ma_list = [50, 60, 70]
ma_column = ['ma_50', 'ma_60', 'ma_70']

# 显示所有列
pd.set_option('display.max_columns', None)
# 显示所有行
pd.set_option('display.max_rows', None)


# 策略回测
def backtesting_ma_bolling(kdata : pd.DataFrame, beginTime=None, endTime=None, tradeDirection=0):
    """
    根据均线排列和布林带回测策略
    入参
    code:股票代码，6位数字代码，此参数不可为空;
    name:股票名称
    beginTime:开始日期（包含），格式“%Y-%m-%d”；
    endTime:结束日期（包含），格式“%Y-%m-%d”；
    tradeDirection:交易方向,0:双向,1:只做多,2:只做空
    """
    # 获取均线排列跟布林带数据
    kdata.index = pd.DatetimeIndex(kdata['日期'])
    calc.df_ema(kdata, field='收盘', ma_list=ma_list)
    calc.df_bolling(kdata, field='收盘')
    # 筛选时间范围
    if beginTime is not None:
        beginT = datetime.strptime(beginTime, "%Y-%m-%d")
        kdata = kdata[kdata.index >= pd.Timestamp(beginT)]
    if endTime is not None:
        endT = datetime.strptime(endTime, "%Y-%m-%d")
        kdata = kdata[kdata.index <= pd.Timestamp(endT)]
    # 交易详情列表,包括进场日期,进场方向,进场价格,止损位置,出场日期,出场价格,持有天数,盈亏比
    trading_list = []

    # 判断是否均线多头
    isHigher = False
    isLower = False
    # 正在布林带上
    isCrossing = False
    # 是否发生了穿过布林带
    isCrossed = False
    # 出现穿过布林带时的多头空头状态,1:多头,-1:空头
    crossMaSequence = 0
    # 发生穿过布林带后,价格偏离过大(离开了布林带,是否能进场交易,0:不进场,1:进场)
    crossDeviateTrade = 1

    currentPosition = {'holding': False}
    for index in kdata.index:
        row = kdata.loc[index]
        # 当前是否有持仓,有持仓判断出场,无持仓判断入场
        if currentPosition['holding'] is True:
            # 判断出场条件, 做多判断触碰布林上轨,做空判断触碰布林下轨
            if currentPosition['direction'] == 'buy':
                # 触碰上轨,以上轨价格平仓
                if row['最低'] <= row['higher_bound'] <= row['最高']:
                    currentPosition['outDate'] = row['日期']
                    currentPosition['outPrice'] = row['higher_bound']
                    profitRatio = round((currentPosition['outPrice'] - currentPosition['entryPrice']) / (currentPosition['entryPrice'] - currentPosition['stop']),2)
                    currentPosition['profitRatio'] = profitRatio
                    currentPosition['isWin'] = True
                    isCrossing = False
                    isCrossed = False
                    isHigher = False
                    isLower = False
                    crossMaSequence = 0
                    trading_list.append(currentPosition)
                    currentPosition = {'holding': False}
                # 止损
                elif row['最低'] < currentPosition['stop']:
                    currentPosition['outDate'] = row['日期']
                    currentPosition['outPrice'] = currentPosition['stop']
                    currentPosition['profitRatio'] = -1
                    currentPosition['isWin'] = False
                    isCrossing = False
                    isCrossed = False
                    isHigher = False
                    isLower = False
                    crossMaSequence = 0
                    trading_list.append(currentPosition)
                    currentPosition = {'holding': False}
            elif currentPosition['direction'] == 'sell':
                # 触碰下轨,以下轨价格平仓
                if row['最低'] <= row['lower_bound'] <= row['最高']:
                    currentPosition['outDate'] = row['日期']
                    currentPosition['outPrice'] = row['lower_bound']
                    profitRatio = round((currentPosition['entryPrice'] - currentPosition['outPrice']) / (currentPosition['stop'] - currentPosition['entryPrice']),2)
                    currentPosition['profitRatio'] = profitRatio
                    currentPosition['isWin'] = True
                    isCrossing = False
                    isCrossed = False
                    isHigher = False
                    isLower = False
                    crossMaSequence = 0
                    trading_list.append(currentPosition)
                    currentPosition = {'holding': False}
                # 止损
                elif row['最高'] > currentPosition['stop']:
                    currentPosition['outDate'] = row['日期']
                    currentPosition['outPrice'] = currentPosition['stop']
                    currentPosition['profitRatio'] = -1
                    currentPosition['isWin'] = False
                    isCrossing = False
                    isCrossed = False
                    isHigher = False
                    isLower = False
                    crossMaSequence = 0
                    trading_list.append(currentPosition)
                    currentPosition = {'holding': False}
        # 止损K可以作为信号K
        if currentPosition['holding'] is False:
        # 止损K不可以作为信号K
        # else:
            # 判断是否满足入场条件
            # 1.均线多头或者空头排列
            maSequence = calc.calc_ma_sequence(row, ma_list, 'ema')
            if (tradeDirection == 0 or tradeDirection == 2) and maSequence == -1:
                isLower = True
                isHigher = False
            if (tradeDirection == 0 or tradeDirection == 1) and maSequence == 1:
                isHigher = True
                isLower = False
            if maSequence == 0:
                isHigher = False
                isLower = False
            # 2.如果之前没有出现穿越布林带,判断是否穿过了布林带的上轨跟下轨(多头下轨,空头上轨)
            if isCrossed is False:
                # 穿过布林带,可能为入场位,记录下止损价
                if isLower:
                    # 判断上穿布林带
                    if row['最低'] <= row['higher_bound'] <= row['最高']:
                        isCrossing = True
                        currentPosition['stop'] = row['最高']
                        isCrossed = True
                        crossMaSequence = -1
                if isHigher:
                    # 判断下穿布林带
                    if row['最低'] <= row['lower_bound'] <= row['最高']:
                        isCrossing = True
                        currentPosition['stop'] = row['最低']
                        isCrossed = True
                        crossMaSequence = 1
            # 3.如果已经出现穿越,则看当前是否是收出了阳线(做多)或者阴线(做空),符合条件就进场
            # 进场条件为多头排列时,价格下穿布林带下轨后收出回到布林带中第一个阳线,空头排列时,价格上
            # 穿布林带上轨后收出回到布林带中第一个阴线
            # 信号K可以作为入场K
            if isCrossed is True:
            # 信号K不可以作为入场K
            # else:
                # 出现了穿过布林带,判断当前是否还是正在布林带上,如果在,更新止损价格,如果回到布林带上并收反向K线进场,如果完全偏离出布林带放弃这次
                if isLower and isCrossing:
                    # 如果多空头方向变了,则放弃
                    if crossMaSequence != -1:
                        isCrossing = False
                        currentPosition['stop'] = 0
                        isCrossed = False
                        crossMaSequence = 0
                    # K线最低价比布林带高,意思是离开了布林带,放弃
                    if crossDeviateTrade == 0 and row['最低'] > row['higher_bound']:
                        isCrossing = False
                        currentPosition['stop'] = 0
                        isCrossed = False
                        crossMaSequence = 0
                    else:
                        # 只要收出阴线进场,阳线继续更新止损价
                        if row['开盘'] <= row['收盘']:
                            currentPosition['stop'] = max(row['最高'], currentPosition['stop'])
                        if row['开盘'] > row['收盘']:
                            # 如果收盘穿过了布林中轨,止损过大,放弃
                            if row['收盘'] > row['mid_bound']:
                                currentPosition['stop'] = max(row['最高'], currentPosition['stop'])
                                currentPosition['holding'] = True
                                currentPosition['entryPrice'] = row['收盘']
                                currentPosition['direction'] = 'sell'
                                currentPosition['entryDate'] = row['日期']
                            else:
                                isCrossing = False
                                currentPosition['stop'] = 0
                                isCrossed = False
                                crossMaSequence = 0
                if isHigher and isCrossing:
                    # 如果多空头方向变了,则放弃
                    if crossMaSequence != 1:
                        isCrossing = False
                        currentPosition['stop'] = 0
                        isCrossed = False
                        crossMaSequence = 0
                    # K线最高价比布林带低,意思是离开了布林带,放弃
                    if crossDeviateTrade == 0 and row['最高'] < row['lower_bound']:
                        isCrossing = False
                        currentPosition['stop'] = 0
                        isCrossed = False
                        crossMaSequence = 0
                    else:
                        # 只要收出阳线进场,阴线继续更新止损价
                        if row['开盘'] >= row['收盘']:
                            currentPosition['stop'] = min(row['最低'], currentPosition['stop'])
                        if row['开盘'] < row['收盘']:
                            # 如果收盘穿过了布林中轨,止损过大,放弃
                            if row['收盘'] < row['mid_bound']:
                                currentPosition['stop'] = min(row['最低'], currentPosition['stop'])
                                currentPosition['holding'] = True
                                currentPosition['entryPrice'] = row['收盘']
                                currentPosition['direction'] = 'buy'
                                currentPosition['entryDate'] = row['日期']
                            else:
                                isCrossing = False
                                currentPosition['stop'] = 0
                                isCrossed = False
                                crossMaSequence = 0
    # 计算总的交易频率跟胜率
    outResult = {}
    if(len(trading_list) <= 0):
        return outResult
    result = pd.DataFrame(trading_list)
    # 总交易次数
    outResult['trade_times'] = len(result)
    # 盈利次数
    outResult['win_times'] = len(result[result['isWin'] == True])
    # 亏损次数
    outResult['lose_times'] = len(result[result['isWin'] == False])
    # 总盈利(盈亏比)
    outResult['profit_ratio'] = round(result['profitRatio'].sum(), 4)
    # 胜率
    outResult['win_chance'] = round(outResult['win_times']/outResult['trade_times'] * 100, 4)
    # 交易频率
    outResult['average_ransaction_requency'] = round(((kdata.tail(1).index - kdata.head(1).index).days / outResult['trade_times']).values[0], 4)
    # 做多次数
    outResult['buy_times'] = len(result[result['direction'] == 'buy'])
    # 做多盈利次数
    outResult['buy_win_times'] = len(result[(result['isWin'] == True) & (result['direction'] == 'buy')])
    # 做多亏损次数
    outResult['buy_lose_times'] = len(result[(result['isWin'] == False) & (result['direction'] == 'buy')])
    # 做空次数
    outResult['sell_times'] = len(result[result['direction'] == 'sell'])
    # 做空盈利次数
    outResult['sell_win_times'] = len(result[(result['isWin'] == True) & (result['direction'] == 'sell')])
    # 做空亏损次数
    outResult['sell_lose_times'] = len(result[(result['isWin'] == False) & (result['direction'] == 'sell')])
    # 最大连续亏损次数
    outResult['max_continuous_lose'] = calc.calc_field_value_times(result, 'isWin', False)

    # 输出数据处理
    result = result.rename(columns={"entryDate": "entryTime", "stop": "stopPrice", "outDate":"outTime"})
    result['profitRatioCumsum'] = result['profitRatio'].cumsum()
    result['holdingTime'] = pd.to_datetime(result['outTime']) - pd.to_datetime(result['entryTime'])
    result['holdingDays'] = result['holdingTime'].map(lambda x: x.days)
    result['holdingHours'] = result['holdingTime'].map(lambda x: x.seconds/3600)
    result.drop('holding', axis='columns', inplace=True)
    result.drop('subgroup', axis='columns', inplace=True)
    result.drop('holdingTime', axis='columns', inplace=True)
    result = result.loc[:, ['entryPrice','direction','stopPrice','outPrice','isWin','profitRatio','profitRatioCumsum','entryTime','outTime','holdingDays','holdingHours']]
    outResult['trade_list'] = result
    return outResult


# 保存策略数据
def save_testing_data(out : {}, code, name='', beginTime=None, endTime=None, excelFileName:str=None):
    if bool(out):
        stdout = '''代码:{}, 名称:{}, 总交易次数:{}, 盈利次数:{}, 亏损次数:{}, 总盈利(盈亏比):{}, 总胜率:{}%, 交易频率:{}/天, 多单次数:{}, 多单盈利次数:{}, 多单亏损次数:{}, 空单次数:{}, 空单盈利次数:{}, 空单亏损次数:{},最大连续亏损次数:{}.'''.format(code, name, out['trade_times'], out['win_times'], out['lose_times'], out['profit_ratio'], out['win_chance'], out['average_ransaction_requency'], out['buy_times'], out['buy_win_times'], out['buy_lose_times'], out['sell_times'], out['sell_win_times'], out['sell_lose_times'], out['max_continuous_lose'])
        logger.info('[' + code + ']' + name + '统计结果:' + stdout)
        # logger.info('[' + code + ']' + name + '交易明细:')
        # logger.info(out['trade_list'])
        # 写入文件
        sheetname = code
        if beginTime is None:
            sheetname = sheetname + '_none'
        else:
            sheetname = sheetname + '_' + beginTime
        if endTime is None:
            sheetname = sheetname + '_none'
        else:
            sheetname = sheetname + '_' + endTime 
        # file = open(os.path.join(scons.ma_bolling_path, sheetname +'_summary.txt'), 'w')
        # file.write(stdout)
        # file.close()
        # out['trade_list'].to_csv(os.path.join(scons.ma_bolling_path, sheetname + ".csv"), encoding="utf-8", index=False)
        saveFileName=''
        ws = None
        if excelFileName is None:
            saveFileName = os.path.join(scons.ma_bolling_path, sheetname + ".xlsx")
        else:
            saveFileName = excelFileName
        #如果文件不存在,先创建
        if os.path.exists(saveFileName) == False:
            wb = Workbook()
            ws = wb.active
            ws.title=sheetname
            wb.save(saveFileName)
            wb = load_workbook(saveFileName)
            ws = wb.get_sheet_by_name(sheetname)
        else:
            wb = load_workbook(saveFileName)
            ws = wb.create_sheet(sheetname)
        # 写入统计信息
        ws.cell(row=1, column=2,value='代码')
        ws.cell(row=1, column=3,value=code)
        ws.cell(row=1, column=4,value='名称')
        ws.cell(row=1, column=5,value=name)
        ws.cell(row=1, column=6,value='总盈利(盈亏比)')
        ws.cell(row=1, column=7,value=out['profit_ratio'])
        ws.cell(row=2, column=2,value='总交易次数')
        ws.cell(row=2, column=3,value=out['trade_times'])
        ws.cell(row=2, column=4,value='盈利次数')
        ws.cell(row=2, column=5,value=out['win_times'])
        ws.cell(row=2, column=6,value='亏损次数')
        ws.cell(row=2, column=7,value=out['lose_times'])
        ws.cell(row=3, column=2,value='总胜率')
        ws.cell(row=3, column=3,value=out['win_chance'])
        ws.cell(row=3, column=4,value='交易频率(天)')
        ws.cell(row=3, column=5,value=out['average_ransaction_requency'])
        ws.cell(row=3, column=6,value='最大连续亏损次数')
        ws.cell(row=3, column=7,value=out['max_continuous_lose'])
        ws.cell(row=4, column=2,value='多单次数')
        ws.cell(row=4, column=3,value=out['buy_times'])
        ws.cell(row=4, column=4,value='多单盈利次数')
        ws.cell(row=4, column=5,value=out['buy_win_times'])
        ws.cell(row=4, column=6,value='多单亏损次数')
        ws.cell(row=4, column=7,value=out['buy_lose_times'])
        ws.cell(row=5, column=2,value='空单次数')
        ws.cell(row=5, column=3,value=out['sell_times'])
        ws.cell(row=5, column=4,value='空单盈利次数')
        ws.cell(row=5, column=5,value=out['sell_win_times'])
        ws.cell(row=5, column=6,value='空单亏损次数')
        ws.cell(row=5, column=7,value=out['sell_lose_times'])

        # 写入交易明细
        ws.append([])
        for r in dataframe_to_rows(out['trade_list'], index=True, header=True):
            if len(r) <= 1:
                continue
            ws.append(r)
        
        # 绘制总盈利(盈亏比)图
        scatter_chart = ScatterChart()
        # 设置标题
        scatter_chart.title = '总盈利(盈亏比)曲线'
        # 设置颜色
        scatter_chart.style = 13
        # 设置x轴y轴标题
        scatter_chart.x_axis.tickLblPos = 'low'
        scatter_chart.width = 30
        scatter_chart.height = 15
        scatter_chart.legend = None
        scatter_chart.x_axis.title = '时间'
        scatter_chart.y_axis.title = '总盈利(盈亏比)'
        
        # 创建x轴的数据来源
        xvalues = Reference(ws, min_col=10, min_row=8, max_row=8 + len(out['trade_list']))
        # 创建yvalues
        yvalues = Reference(ws, min_col=8, min_row=7, max_row=7 + len(out['trade_list']))
        series = Series(yvalues, xvalues=xvalues, title_from_data=True)
        scatter_chart.series.append(series)
        # 将散点图添加到ws工作表中
        ws.add_chart(scatter_chart, 'N1')
        wb.save(saveFileName)


# 单个外汇回测
def backtesting_forex(code, cycle, name='', beginTime=None, endTime=None, tradeDirection=0, excelFileName:str=None):
    """
    外汇回测
    入参
    code:外汇代码,大写
    cycle:外汇周期(H1, H4, D1)
    beginTime:开始日期（包含），格式“%Y-%m-%d”
    endTime:结束日期（包含），格式“%Y-%m-%d”
    tradeDirection:交易方向,0:双向,1:只做多,2:只做空
    """
    filename = code + '_' + cycle
    logger.info('[' + filename + ']' + '回测开始')
    file_name = os.path.join(ccons.forex_history_path, filename+".csv")
    exist = os.path.exists(file_name)
    if not exist:
        logger.info(file_name + '数据文件不存在.回测结束')
        return
    kdata = pd.read_csv(file_name, sep=r'\t')
    # 表头格式化
    if cycle == 'D1':
        kdata = kdata.rename(columns={"<DATE>": "日期"})
    else:
        kdata['日期'] = pd.to_datetime(kdata.apply(lambda x: str(x['<DATE>']) + " " + str(x['<TIME>']), axis=1), format="%Y.%m.%d %H:%M:%S")
    kdata['日期'] = pd.to_datetime(kdata['日期'])
    kdata = kdata.rename(columns={"<OPEN>":"开盘", "<HIGH>":"最高", "<LOW>":"最低", "<CLOSE>":"收盘"})
    out = backtesting_ma_bolling(kdata, beginTime, endTime, tradeDirection)
    save_testing_data(out, filename, name, beginTime=beginTime, endTime=endTime, excelFileName=excelFileName)
    logger.info('[' + filename + ']' + '回测结束')
    out['code'] = filename
    out['name'] = name
    return out


# 所有外汇回测
def backtesting_all_forex(cycle:[], beginTime=None, endTime=None, tradeDirection=0):
    basic = pd.read_csv(ccons.forex_basic_file, dtype={'代码': str})
    summaryList = []
    # 先创建excel文件
    filename = os.path.join(scons.ma_bolling_path, 'all_forex_summary_' + datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx')
    for index in basic.index:
        row = basic.loc[index, :]
        s_code = row['代码']
        name = row['名称']
        for cyc in cycle:
            try:
                out = backtesting_forex(s_code, cyc, name, beginTime, endTime, tradeDirection, filename)
                if bool(out):
                    # stdout = '''总交易次数:{}, 盈利次数:{}, 亏损次数:{}, 总盈利(盈亏比):{}, 总胜率:{}%, 交易频率:{}/天, 多单次数:{}, 多单盈利次数:{}, 多单亏损次数:{}, 空单次数:{}, 空单盈利次数:{}, 空单亏损次数:{},最大连续亏损次数:{}.'''.format(out['trade_times'], out['win_times'], out['lose_times'], out['profit_ratio'], out['win_chance'], out['average_ransaction_requency'], out['buy_times'], out['buy_win_times'], out['buy_lose_times'], out['sell_times'], out['sell_win_times'], out['sell_lose_times'], out['max_continuous_lose'])
                    # logger.info('[' + out['code'] + ']' + out['name'] + '回测完成,统计结果:' + stdout)
                    # logger.info('[' + out['code'] + ']' + out['name'] + '交易明细:')
                    # logger.info(out['trade_list'])
                    del out['trade_list']
                    summaryList.append(out)
            except BaseException:
                logger.error(str(s_code) + ':回测失败,原因:' + traceback.format_exc())
    summary = pd.DataFrame(summaryList)
    summaryOut = '''总计,交易次数:{}, 盈利次数:{}, 亏损次数:{},  总盈利(盈亏比):{}, 总胜率:{}%'''.format(summary['trade_times'].sum(), summary['win_times'].sum(), summary['lose_times'].sum(), summary['profit_ratio'].sum(), round(summary['win_times'].sum()/summary['trade_times'].sum() * 100, 4))
    logger.info(summaryOut)
    # file = open(os.path.join(scons.ma_bolling_path, 'all_forex_summary' + datetime.now().strftime('%Y%m%d%H%M%S') + '.txt'), 'w')
    # file.write(summaryOut)
    # file.close()
    # summary.to_csv(os.path.join(scons.ma_bolling_path, 'all_forex_summary_list' + datetime.now().strftime('%Y%m%d%H%M%S') + '.csv'), encoding="utf-8", index=False)
    # 保存excel
    wb = load_workbook(filename)
    ws = wb.create_sheet('all_forex_summary', 0)
    # 写入统计明细
    for r in dataframe_to_rows(summary, index=True, header=True):
        if len(r) <= 1:
            continue
        ws.append(r)
    # 统计信息
    maxRow = ws.max_row
    ws['A'+str(maxRow+1)].value = "总计"
    ws['B'+str(maxRow+1)].value = "=SUM(B2:B"+str(maxRow)+")"
    ws['C'+str(maxRow+1)].value = "=SUM(C2:C"+str(maxRow)+")"
    ws['D'+str(maxRow+1)].value = "=SUM(D2:D"+str(maxRow)+")"
    ws['E'+str(maxRow+1)].value = "=SUM(E2:E"+str(maxRow)+")"
    ws['F'+str(maxRow+1)].value = "=C"+str(maxRow+1)+"/B"+str(maxRow+1)+"*100"
    wb.save(filename)


# 单个股票回测
def backtesting_stock(code, name='', beginTime=None, endTime=None, tradeDirection=0, excelFileName:str=None):
    logger.info('[' + code + ']' + name + '回测开始')
    file_name = os.path.join(ccons.stock_history_path, code+".csv")
    exist = os.path.exists(file_name)
    if not exist:
        logger.info(file_name + '数据文件不存在.回测结束')
        return
    kdata = pd.read_csv(file_name)
    # 格式化日期
    kdata['日期'] = pd.to_datetime(kdata.apply(lambda x: str(x['日期']), axis=1), format="%Y-%m-%d")
    kdata['日期'] = kdata['日期'].dt.strftime('%Y-%m-%d %H:%M:%S')
    out = backtesting_ma_bolling(kdata, beginTime, endTime, tradeDirection)
    save_testing_data(out, code, name, beginTime=beginTime, endTime=endTime, excelFileName=excelFileName)
    logger.info('[' + code + ']' + name + '回测结束')
    out['code'] = code
    out['name'] = name
    return out


# 所有股票回测
def backtesting_all_stock(beginTime=None, endTime=None, tradeDirection=0):
    basic = pd.read_csv(ccons.stock_basic_file, dtype={'代码': str})
    summaryList = []
    # 先创建excel文件
    filename = os.path.join(scons.ma_bolling_path, 'all_stock_summary_' + datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx')
    for index in basic.index:
        row = basic.loc[index, :]
        s_code = row['代码']
        name = row['名称']
        try:
            out = backtesting_stock(s_code, name, beginTime, endTime, tradeDirection)
            if bool(out):
                # stdout = '''总交易次数:{}, 盈利次数:{}, 亏损次数:{}, 总盈利(盈亏比):{}, 总胜率:{}%, 交易频率:{}/天, 多单次数:{}, 多单盈利次数:{}, 多单亏损次数:{}, 空单次数:{}, 空单盈利次数:{}, 空单亏损次数:{},最大连续亏损次数:{}.'''.format(out['trade_times'], out['win_times'], out['lose_times'], out['profit_ratio'], out['win_chance'], out['average_ransaction_requency'], out['buy_times'], out['buy_win_times'], out['buy_lose_times'], out['sell_times'], out['sell_win_times'], out['sell_lose_times'], out['max_continuous_lose'])
                # logger.info('[' + out['code'] + ']' + out['name'] + '回测完成,统计结果:' + stdout)
                # logger.info('[' + out['code'] + ']' + out['name'] + '交易明细:')
                # logger.info(out['trade_list'])
                del out['trade_list']
                summaryList.append(out)
        except BaseException:
            logger.error(str(s_code) + ':回测失败,原因:' + traceback.format_exc())
    summary = pd.DataFrame(summaryList)
    summaryOut = '''总计,交易次数:{}, 盈利次数:{}, 亏损次数:{},  总盈利(盈亏比):{}, 总胜率:{}%'''.format(summary['trade_times'].sum(), summary['win_times'].sum(), summary['lose_times'].sum(), summary['profit_ratio'].sum(), round(summary['win_times'].sum()/summary['trade_times'].sum() * 100, 4))
    logger.info(summaryOut)
    # file = open(os.path.join(scons.ma_bolling_path, 'all_stock_summary' + datetime.now().strftime('%Y%m%d%H%M%S') + '.txt'), 'w')
    # file.write(summaryOut)
    # file.close()
    # summary.to_csv(os.path.join(scons.ma_bolling_path, 'all_stock_summary_list' + datetime.now().strftime('%Y%m%d%H%M%S') + '.csv'), encoding="utf-8", index=False)
    # 保存excel
    wb = load_workbook(filename)
    ws = wb.create_sheet('all_stock_summary', 0)
    # 写入统计明细
    for r in dataframe_to_rows(summary, index=True, header=True):
        if len(r) <= 1:
            continue
        ws.append(r)
    # 统计信息
    maxRow = ws.max_row
    ws['A'+str(maxRow+1)].value = "总计"
    ws['B'+str(maxRow+1)].value = "=SUM(B2:B"+str(maxRow)+")"
    ws['C'+str(maxRow+1)].value = "=SUM(C2:C"+str(maxRow)+")"
    ws['D'+str(maxRow+1)].value = "=SUM(D2:D"+str(maxRow)+")"
    ws['E'+str(maxRow+1)].value = "=SUM(E2:E"+str(maxRow)+")"
    ws['F'+str(maxRow+1)].value = "=C"+str(maxRow+1)+"/B"+str(maxRow+1)+"*100"
    wb.save(filename)


if __name__ == '__main__':
    # backtesting_all_stock(tradeDirection=1)
    # backtesting_stock('000002', '万科A')
    # backtesting_forex('EURUSD','H4', beginTime='2022-01-01', endTime='2023-09-20')
    backtesting_all_forex(['H1'], beginTime='2022-01-01', endTime='2022-12-31')
